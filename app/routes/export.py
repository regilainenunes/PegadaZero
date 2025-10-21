from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import Response
from typing import Literal
from datetime import datetime

from sqlalchemy.orm import Session
from ..sql_models import UserSQL, CarbonRecordSQL
from ..models.carbon import Category
from ..utils.auth_handler import get_current_user
from ..utils.sql_session import get_db

from jinja2 import Environment, FileSystemLoader, select_autoescape
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter


router = APIRouter()


env = Environment(
    loader=FileSystemLoader("app/templates"),
    autoescape=select_autoescape(["html", "xml"]),
)


@router.get("/pdf")
async def export_pdf(current_user: UserSQL = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        from weasyprint import HTML  # import lazy para evitar falha ao inicializar sem a lib
    except Exception:
        raise HTTPException(status_code=501, detail="PDF indisponível: WeasyPrint não instalado")
    records = db.query(CarbonRecordSQL).filter(CarbonRecordSQL.user_id == current_user.id).all()
    totals = {c.value: 0.0 for c in Category}
    for r in records:
        totals[r.category] += r.emissions

    template = env.get_template("report.html")
    html_str = template.render(
        title="Relatório de Emissões - PegadaZero",
        generated_at=datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC"),
        user=current_user,
        totals=totals,
        records=records,
    )

    pdf_bytes = HTML(string=html_str).write_pdf()
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=relatorio_pegadazero.pdf"
        },
    )


@router.get("/excel")
async def export_excel(
    fmt: Literal["xlsx", "xlsm"] = Query("xlsx"),
    current_user: UserSQL = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    records = db.query(CarbonRecordSQL).filter(CarbonRecordSQL.user_id == current_user.id).all()
    totals = {c.value: 0.0 for c in Category}
    for r in records:
        totals[r.category] += r.emissions

    wb = Workbook()
    if fmt == "xlsm":
        wb.template = True  # salva como arquivo habilitado a macro (sem macros)
    ws = wb.active
    ws.title = "Resumo"

    # Estilos
    header_fill = PatternFill("solid", fgColor="22C55E")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Cabeçalho
    ws.append(["Categoria", "Emissões (kg CO2e)"])
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # Dados
    row = 2
    for cat, value in totals.items():
        ws.append([cat.capitalize(), value])
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).number_format = "0.000"
        ws.cell(row=row, column=2).border = thin_border
        row += 1

    # Largura das colunas
    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 22

    # Segunda aba com registros
    ws2 = wb.create_sheet(title="Registros")
    ws2.append(["Data", "Categoria", "Quantidade", "Emissões (kg CO2e)"])
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    r = 2
    for rec in records:
        ws2.append([
            rec.date.strftime("%Y-%m-%d"),
            rec.category.capitalize(),
            rec.amount,
            rec.emissions,
        ])
        ws2.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws2.cell(row=r, column=3).number_format = "0.000"
        ws2.cell(row=r, column=4).number_format = "0.000"
        for c in range(1, 5):
            ws2.cell(row=r, column=c).border = thin_border
        r += 1

    ws2.column_dimensions[get_column_letter(1)].width = 16
    ws2.column_dimensions[get_column_letter(2)].width = 18
    ws2.column_dimensions[get_column_letter(3)].width = 18
    ws2.column_dimensions[get_column_letter(4)].width = 22

    # Salvar em memória
    bio = BytesIO()
    if fmt == "xlsm":
        wb.save(bio)
        content_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
        filename = "relatorio_pegadazero.xlsm"
    else:
        wb.save(bio)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "relatorio_pegadazero.xlsx"
    bio.seek(0)

    return Response(
        content=bio.read(),
        media_type=content_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        },
    )